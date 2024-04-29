import os
import openpyxl

# Set the path to the main folder
main_folder = "/Users/qichenyuan/Desktop/Synthego Data Summary/CRISPRessoBatch_on_96plex100pmol"
# Set the path for the Excel file
excel_file_path = "/Users/qichenyuan/Desktop/Synthego Data Summary/Analysis.xlsx"
# Workbook
workbook = openpyxl.load_workbook(excel_file_path)


# Function to check the conditions and update the Excel file
def Perfect_Editing(file_path, worksheet, partial_name):
    try:
        with open(file_path, 'r') as file:
            lines = file.readlines()
            headers = lines[0].strip().split('\t')

            aligned_sequence_index = headers.index("Aligned_Sequence")
            reference_sequence_index = headers.index("Reference_Sequence")
            percent_reads_index = headers.index("%Reads")

            # Flag to track if any condition matches
            condition_matched = False

            for line in lines[1:]:
                data = line.strip().split('\t')
                reference_sequence = data[reference_sequence_index]
                aligned_sequence = data[aligned_sequence_index]
                percent_reads = float(data[percent_reads_index])  # Convert to float

                # Define conditions for matching within the protospacer range (10,33)
                condition_1 = (
                        reference_sequence[14] == "A" and aligned_sequence[14] == "G" and
                        reference_sequence[31] == aligned_sequence[31] == "G" and
                        reference_sequence[32] == aligned_sequence[32] == "G" and
                        all(reference_sequence[i] == aligned_sequence[i] for i in range(10, 33) if
                            i not in [14, 31, 32])
                )

                condition_2 = (
                        reference_sequence[15] == "A" and aligned_sequence[15] == "G" and
                        reference_sequence[31] == aligned_sequence[31] == "G" and
                        reference_sequence[32] == aligned_sequence[32] == "G" and
                        all(reference_sequence[i] == aligned_sequence[i] for i in range(10, 33) if
                            i not in [15, 31, 32])
                )
                # be extremely careful that the index of number 25 base is 24
                condition_3 = (
                        reference_sequence[24] == "T" and aligned_sequence[24] == "C" and
                        reference_sequence[7] == aligned_sequence[7] == "C" and
                        reference_sequence[8] == aligned_sequence[8] == "C" and
                        all(reference_sequence[i] == aligned_sequence[i] for i in range(10, 33) if i not in [24, 7, 8])
                )
                # be extremely careful that the index of number 26 bas is 25
                condition_4 = (
                        reference_sequence[25] == "T" and aligned_sequence[25] == "C" and
                        reference_sequence[7] == aligned_sequence[7] == "C" and
                        reference_sequence[8] == aligned_sequence[8] == "C" and
                        all(reference_sequence[i] == aligned_sequence[i] for i in range(10, 33) if i not in [25, 7, 8])
                )
                # Check conditions and append to worksheet
                if condition_1 or condition_2 or condition_3 or condition_4:
                    worksheet.append({"A": partial_name, "B": percent_reads})
                    condition_matched = True
                    break  # No need to check further conditions

            # If no condition matched, write N/A
            if not condition_matched:
                worksheet.append({"A": partial_name, "B": "N/A"})

    except Exception as e:
        print(f"Error processing file {file_path}: {e}")


# Create a new worksheet with the extracted sheet name
worksheet = workbook.create_sheet(title='_'.join(['Perfect', 'Editing']))

# Set headers in columns A and B
worksheet['A1'] = 'Name'
worksheet['B1'] = 'Perfect Editing %'

# Iterate through subfolders
for folder_name in os.listdir(main_folder):
    if "CRISPResso_on_" in folder_name:
        subfolder_path = os.path.join(main_folder, folder_name)

        # Check if the item is a directory before processing
        if os.path.isdir(subfolder_path):
            partial_name = folder_name.replace("CRISPResso_on_", "")
            text_file_name = [file for file in os.listdir(subfolder_path) if
                              "Alleles_frequency_table_around_sgRNA_" in file and file.endswith(".txt")]

            if text_file_name:
                text_file_path = os.path.join(subfolder_path, text_file_name[0])
                Perfect_Editing(text_file_path, worksheet, partial_name)

# Save the Excel workbook
workbook.save(excel_file_path)


# Function to check the conditions and update the Excel file
def Perfect_And_Bystander_Editing(file_path, worksheet, partial_name):
    try:
        with open(file_path, 'r') as file:
            lines = file.readlines()
            headers = lines[0].strip().split('\t')

            aligned_sequence_index = headers.index("Aligned_Sequence")
            reference_sequence_index = headers.index("Reference_Sequence")
            percent_reads_index = headers.index("%Reads")

            # Variables to accumulate the percentage reads
            total_percent_reads = 0

            for line in lines[1:]:
                data = line.strip().split('\t')
                reference_sequence = data[reference_sequence_index]
                aligned_sequence = data[aligned_sequence_index]
                percent_reads = float(data[percent_reads_index])  # Convert to float

                # Define conditions for matching
                condition_1 = (
                        reference_sequence[14] == "A" and aligned_sequence[14] == "G" and
                        reference_sequence[31] == aligned_sequence[31] == "G" and
                        reference_sequence[32] == aligned_sequence[32] == "G")

                condition_2 = (
                        reference_sequence[15] == "A" and aligned_sequence[15] == "G" and
                        reference_sequence[31] == aligned_sequence[31] == "G" and
                        reference_sequence[32] == aligned_sequence[32] == "G")

                condition_3 = (
                        reference_sequence[24] == "T" and aligned_sequence[24] == "C" and
                        reference_sequence[7] == aligned_sequence[7] == "C" and
                        reference_sequence[8] == aligned_sequence[8] == "C")

                condition_4 = (
                        reference_sequence[25] == "T" and aligned_sequence[25] == "C" and
                        reference_sequence[7] == aligned_sequence[7] == "C" and
                        reference_sequence[8] == aligned_sequence[8] == "C")

                # Check conditions and accumulate percentage reads
                if condition_1 or condition_2 or condition_3 or condition_4:
                    total_percent_reads += percent_reads

            # Write total percentage reads to the worksheet
            if total_percent_reads > 0:
                worksheet.append({"A": partial_name, "B": total_percent_reads})
            else:
                worksheet.append({"A": partial_name, "B": "N/A"})  # No matches, write N/A

    except Exception as e:
        print(f"Error processing file {file_path}: {e}")


# Create a new worksheet with the extracted sheet name
worksheet = workbook.create_sheet(title='_'.join(['Perfect', 'And', 'Bystander', 'Editing']))

# Set headers in columns A and B
worksheet['A1'] = 'Name'
worksheet['B1'] = 'Perfect and Bystander Editing%'

# Iterate through subfolders
for folder_name in os.listdir(main_folder):
    if "CRISPResso_on_" in folder_name:
        subfolder_path = os.path.join(main_folder, folder_name)

        # Check if the item is a directory before processing
        if os.path.isdir(subfolder_path):
            partial_name = folder_name.replace("CRISPResso_on_", "")
            text_file_name = [file for file in os.listdir(subfolder_path) if
                              "Alleles_frequency_table_around_sgRNA_" in file and file.endswith(".txt")]

            if text_file_name:
                text_file_path = os.path.join(subfolder_path, text_file_name[0])
                Perfect_And_Bystander_Editing(text_file_path, worksheet, partial_name)

# Save the Excel workbook
workbook.save(excel_file_path)


# Function to check the conditions and update the Excel file
def Wild_Type_Percentage(file_path, worksheet, partial_name):
    try:
        with open(file_path, 'r') as file:
            lines = file.readlines()
            headers = lines[0].strip().split('\t')

            aligned_sequence_index = headers.index("Aligned_Sequence")
            reference_sequence_index = headers.index("Reference_Sequence")
            percent_reads_index = headers.index("%Reads")

            # Flag to track if any condition matches
            condition_matched = False

            for line in lines[1:]:
                data = line.strip().split('\t')
                reference_sequence = data[reference_sequence_index]
                aligned_sequence = data[aligned_sequence_index]
                percent_reads = float(data[percent_reads_index])  # Convert to float

                # Define conditions for matching
                condition_WT = (
                    all(reference_sequence[i] == aligned_sequence[i] for i in range(10, 33))
                )

                # Check conditions and append to worksheet
                if condition_WT:
                    worksheet.append({"A": partial_name, "B": percent_reads})
                    condition_matched = True
                    break  # No need to check further conditions

            # If no condition matched, write N/A
            if not condition_matched:
                worksheet.append({"A": partial_name, "B": "N/A"})

    except Exception as e:
        print(f"Error processing file {file_path}: {e}")


# Create a new worksheet with the extracted sheet name
worksheet = workbook.create_sheet(title='_'.join(['Wild', 'Type', 'Percentage']))

# Set headers in columns A and B
worksheet['A1'] = 'Name'
worksheet['B1'] = 'Wild-Type %'

# Iterate through subfolders
for folder_name in os.listdir(main_folder):
    if "CRISPResso_on_" in folder_name:
        subfolder_path = os.path.join(main_folder, folder_name)

        # Check if the item is a directory before processing
        if os.path.isdir(subfolder_path):
            partial_name = folder_name.replace("CRISPResso_on_", "")
            text_file_name = [file for file in os.listdir(subfolder_path) if
                              "Alleles_frequency_table_around_sgRNA_" in file and file.endswith(".txt")]

            if text_file_name:
                text_file_path = os.path.join(subfolder_path, text_file_name[0])
                Wild_Type_Percentage(text_file_path, worksheet, partial_name)

# Save the Excel workbook
workbook.save(excel_file_path)
